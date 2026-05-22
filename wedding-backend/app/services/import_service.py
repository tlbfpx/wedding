from __future__ import annotations

import csv
import io
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Customer, Supplier
from app.models.customer import CustomerStatus, Gender
from app.models.supplier import SupplierType, CooperationStatus
from app.schemas.import_schema import ImportResult
from app.utils.errors import AppException

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
MAX_ROW_COUNT = 1000

CUSTOMER_TEMPLATE_HEADERS = ["姓名", "手机号", "性别", "来源", "预算范围", "婚期", "备注"]
SUPPLIER_TEMPLATE_HEADERS = ["名称", "类型", "联系人", "联系电话", "地址", "合作状态", "备注"]

SUPPLIER_TYPE_MAP = {
    "四大金刚": "four_gods",
    "婚车": "car",
    "场地": "venue",
    "花艺": "floral",
    "摄影": "photo",
    "主持": "host",
    "其他": "other",
}

COOPERATION_STATUS_MAP = {
    "合作中": "active",
    "已暂停": "suspended",
    "已拉黑": "blacklist",
}

VALID_GENDERS = {"男": "male", "女": "female"}

PHONE_PATTERN = r"^1[3-9]\d{9}$"


class ImportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_template(self, import_type: str) -> StreamingResponse:
        if import_type == "customer":
            headers = CUSTOMER_TEMPLATE_HEADERS
            sheet_title = "客户导入模板"
            filename = "customer_template.xlsx"
        elif import_type == "supplier":
            headers = SUPPLIER_TEMPLATE_HEADERS
            sheet_title = "供应商导入模板"
            filename = "supplier_template.xlsx"
        else:
            raise AppException(400, "INVALID_TYPE", f"不支持的导入类型: {import_type}")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    async def import_customers(self, file) -> ImportResult:
        content = await file.read()

        if len(content) > MAX_FILE_SIZE:
            raise AppException(400, "FILE_TOO_LARGE", "文件大小不能超过5MB")

        filename = file.filename or ""
        rows = self._parse_file(content, filename)

        if len(rows) > MAX_ROW_COUNT:
            raise AppException(400, "TOO_MANY_ROWS", "数据行数不能超过1000行")

        # Validate headers
        if not rows or rows[0] != CUSTOMER_TEMPLATE_HEADERS:
            raise AppException(400, "INVALID_HEADERS", "表头不匹配，请使用标准导入模板")

        errors = []
        valid_rows = []
        seen_phones = set()

        # Collect all phones from DB for uniqueness check
        all_phones_result = await self.db.execute(select(Customer.phone))
        existing_phones = {r for r in all_phones_result.scalars().all()}

        import re
        phone_re = re.compile(PHONE_PATTERN)

        for idx, row in enumerate(rows[1:], start=2):  # start from 2 (1-indexed, skip header)
            row_errors = []

            # Pad row if shorter than headers
            padded = list(row) + [""] * (len(CUSTOMER_TEMPLATE_HEADERS) - len(row))

            name = str(padded[0]).strip()
            phone = str(padded[1]).strip()
            gender_str = str(padded[2]).strip()
            # source, budget_range, wedding_date, note
            budget_range = str(padded[4]).strip()
            wedding_date_str = str(padded[5]).strip()
            note = str(padded[6]).strip()

            # Name required
            if not name:
                row_errors.append({"row": idx, "field": "姓名", "message": "姓名不能为空"})

            # Phone required + format check
            if not phone:
                row_errors.append({"row": idx, "field": "手机号", "message": "手机号不能为空"})
            elif not phone_re.match(phone):
                row_errors.append({"row": idx, "field": "手机号", "message": "手机号格式不正确"})
            elif phone in existing_phones:
                row_errors.append({"row": idx, "field": "手机号", "message": "手机号已存在"})
            elif phone in seen_phones:
                row_errors.append({"row": idx, "field": "手机号", "message": "文件内手机号重复"})

            # Gender validation
            gender_value = "unknown"
            if gender_str:
                if gender_str in VALID_GENDERS:
                    gender_value = VALID_GENDERS[gender_str]
                else:
                    row_errors.append({"row": idx, "field": "性别", "message": "性别必须为 男 或 女"})

            if row_errors:
                errors.extend(row_errors)
                continue

            seen_phones.add(phone)
            existing_phones.add(phone)

            # Parse wedding_date
            wedding_date = None
            if wedding_date_str:
                try:
                    wedding_date = datetime.strptime(wedding_date_str, "%Y-%m-%d").date()
                except ValueError:
                    row_errors.append({"row": idx, "field": "婚期", "message": "婚期格式不正确，请使用 YYYY-MM-DD"})
                    errors.extend(row_errors)
                    continue

            valid_rows.append({
                "name": name,
                "phone": phone,
                "gender": gender_value,
                "budget_range": budget_range or None,
                "wedding_date": wedding_date,
                "note": note or None,
            })

        # Batch insert
        for row_data in valid_rows:
            customer = Customer(
                name=row_data["name"],
                phone=row_data["phone"],
                gender=row_data["gender"],
                status=CustomerStatus.potential,
                budget_range=row_data["budget_range"],
                wedding_date=row_data["wedding_date"],
                note=row_data["note"],
                assigned_sale_id=None,
            )
            self.db.add(customer)

        await self.db.commit()

        return ImportResult(
            total=len(rows) - 1,
            success=len(valid_rows),
            failed=len(rows) - 1 - len(valid_rows),
            errors=errors,
        )

    async def import_suppliers(self, file) -> ImportResult:
        content = await file.read()

        if len(content) > MAX_FILE_SIZE:
            raise AppException(400, "FILE_TOO_LARGE", "文件大小不能超过5MB")

        filename = file.filename or ""
        rows = self._parse_file(content, filename)

        if len(rows) > MAX_ROW_COUNT:
            raise AppException(400, "TOO_MANY_ROWS", "数据行数不能超过1000行")

        # Validate headers
        if not rows or rows[0] != SUPPLIER_TEMPLATE_HEADERS:
            raise AppException(400, "INVALID_HEADERS", "表头不匹配，请使用标准导入模板")

        errors = []
        valid_rows = []

        for idx, row in enumerate(rows[1:], start=2):
            row_errors = []

            # Pad row if shorter than headers
            padded = list(row) + [""] * (len(SUPPLIER_TEMPLATE_HEADERS) - len(row))

            name = str(padded[0]).strip()
            type_str = str(padded[1]).strip()
            contact = str(padded[2]).strip()
            phone = str(padded[3]).strip()
            address = str(padded[4]).strip()
            cooperation_status_str = str(padded[5]).strip()
            note = str(padded[6]).strip()

            # Name required
            if not name:
                row_errors.append({"row": idx, "field": "名称", "message": "名称不能为空"})

            # Type validation
            if not type_str:
                row_errors.append({"row": idx, "field": "类型", "message": "类型不能为空"})
            elif type_str not in SUPPLIER_TYPE_MAP:
                row_errors.append({
                    "row": idx,
                    "field": "类型",
                    "message": f"类型必须为: {', '.join(SUPPLIER_TYPE_MAP.keys())}",
                })

            # Cooperation status validation (optional, defaults to active)
            cooperation_status_value = CooperationStatus.active
            if cooperation_status_str:
                if cooperation_status_str in COOPERATION_STATUS_MAP:
                    cooperation_status_value = CooperationStatus(COOPERATION_STATUS_MAP[cooperation_status_str])
                else:
                    row_errors.append({
                        "row": idx,
                        "field": "合作状态",
                        "message": f"合作状态必须为: {', '.join(COOPERATION_STATUS_MAP.keys())}",
                    })

            if row_errors:
                errors.extend(row_errors)
                continue

            valid_rows.append({
                "name": name,
                "type": SupplierType(SUPPLIER_TYPE_MAP[type_str]),
                "contact": contact or None,
                "phone": phone or None,
                "address": address or None,
                "cooperation_status": cooperation_status_value,
                "rating": None,
                "note": note or None,
            })

        # Batch insert
        for row_data in valid_rows:
            supplier = Supplier(
                name=row_data["name"],
                type=row_data["type"],
                contact=row_data["contact"],
                phone=row_data["phone"],
                address=row_data["address"],
                cooperation_status=row_data["cooperation_status"],
                rating=row_data["rating"] or 0.0,
                note=row_data["note"],
            )
            self.db.add(supplier)

        await self.db.commit()

        return ImportResult(
            total=len(rows) - 1,
            success=len(valid_rows),
            failed=len(rows) - 1 - len(valid_rows),
            errors=errors,
        )

    def _parse_file(self, content: bytes, filename: str) -> list[list[str]]:
        if filename.endswith(".xlsx"):
            return self._parse_xlsx(content)
        elif filename.endswith(".csv"):
            return self._parse_csv(content)
        else:
            raise AppException(400, "INVALID_FORMAT", "仅支持 xlsx 或 csv 格式文件")

    def _parse_xlsx(self, content: bytes) -> list[list[str]]:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()
        return rows

    def _parse_csv(self, content: bytes) -> list[list[str]]:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        return [row for row in reader]
