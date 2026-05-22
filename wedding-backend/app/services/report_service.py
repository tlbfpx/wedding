from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Order, Customer, Payment, User
from app.models.customer import FollowUp
from app.utils.errors import AppException


ORDER_STATUS_MAP = {
    "intention": "意向",
    "signed": "已签约",
    "executing": "执行中",
    "completed": "已完成",
    "cancelled": "已取消",
}

CUSTOMER_STATUS_MAP = {
    "potential": "潜在",
    "following": "跟进中",
    "intention": "有意向",
    "signed": "已签约",
    "lost": "已流失",
}

PAYMENT_METHOD_MAP = {
    "cash": "现金",
    "transfer": "转账",
    "wechat": "微信",
    "alipay": "支付宝",
    "card": "刷卡",
}

ORDER_HEADERS = ["订单编号", "客户姓名", "销售", "策划师", "订单状态", "订单总额", "已付金额", "折扣", "创建时间"]
CUSTOMER_HEADERS = ["客户姓名", "手机号", "状态", "来源", "预算范围", "婚期", "负责销售", "跟进次数", "最近跟进", "创建时间"]
FINANCE_HEADERS = ["订单编号", "客户姓名", "付款方式", "付款金额", "付款时间", "订单总额", "已付总额", "待付金额"]

MAX_EXPORT_ROWS = 10000


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_orders(
        self,
        date_start: Optional[str] = None,
        date_end: Optional[str] = None,
        status: Optional[str] = None,
        sale_id: Optional[int] = None,
    ) -> StreamingResponse:
        query = (
            select(Order, Customer.name.label("customer_name"))
            .join(Customer, Order.customer_id == Customer.id, isouter=True)
        )

        if date_start:
            query = query.where(Order.created_at >= datetime.fromisoformat(date_start))
        if date_end:
            query = query.where(Order.created_at <= datetime.fromisoformat(date_end))
        if status:
            query = query.where(Order.status == status)
        if sale_id:
            query = query.where(Order.sale_id == sale_id)

        query = query.order_by(Order.created_at.desc()).limit(MAX_EXPORT_ROWS)
        result = await self.db.execute(query)
        rows = result.all()

        # Fetch sale and planner names in batch
        user_ids = set()
        for row in rows:
            order = row[0]
            user_ids.add(order.sale_id)
            if order.planner_id:
                user_ids.add(order.planner_id)

        user_map = {}
        if user_ids:
            user_result = await self.db.execute(
                select(User).where(User.id.in_(user_ids))
            )
            for user in user_result.scalars().all():
                user_map[user.id] = user.name

        wb = Workbook()
        ws = wb.active
        ws.title = "订单报表"
        self._write_header(ws, ORDER_HEADERS)

        for row_data in rows:
            order = row_data[0]
            customer_name = row_data[1]
            ws.append([
                order.order_no,
                customer_name or "",
                user_map.get(order.sale_id, ""),
                user_map.get(order.planner_id, "") if order.planner_id else "",
                ORDER_STATUS_MAP.get(order.status.value if hasattr(order.status, "value") else str(order.status), str(order.status)),
                float(order.total_amount),
                float(order.paid_amount),
                float(order.discount),
                order.created_at.strftime("%Y-%m-%d %H:%M:%S") if order.created_at else "",
            ])

        return self._build_response(wb, "order")

    async def export_customers(
        self,
        date_start: Optional[str] = None,
        date_end: Optional[str] = None,
        status: Optional[str] = None,
        sale_id: Optional[int] = None,
    ) -> StreamingResponse:
        query = select(Customer)

        if date_start:
            query = query.where(Customer.created_at >= datetime.fromisoformat(date_start))
        if date_end:
            query = query.where(Customer.created_at <= datetime.fromisoformat(date_end))
        if status:
            query = query.where(Customer.status == status)
        if sale_id:
            query = query.where(Customer.assigned_sale_id == sale_id)

        query = query.order_by(Customer.created_at.desc()).limit(MAX_EXPORT_ROWS)
        result = await self.db.execute(query)
        customers = result.scalars().all()

        # Batch fetch sale names
        sale_ids = {c.assigned_sale_id for c in customers if c.assigned_sale_id}
        user_map = {}
        if sale_ids:
            user_result = await self.db.execute(
                select(User).where(User.id.in_(sale_ids))
            )
            for user in user_result.scalars().all():
                user_map[user.id] = user.name

        # Batch fetch follow-up stats
        customer_ids = [c.id for c in customers]
        follow_up_map = {}
        last_follow_map = {}
        if customer_ids:
            fu_count_result = await self.db.execute(
                select(FollowUp.customer_id, func.count(FollowUp.id))
                .where(FollowUp.customer_id.in_(customer_ids))
                .group_by(FollowUp.customer_id)
            )
            for cid, cnt in fu_count_result.all():
                follow_up_map[cid] = cnt

            fu_last_result = await self.db.execute(
                select(FollowUp.customer_id, func.max(FollowUp.created_at))
                .where(FollowUp.customer_id.in_(customer_ids))
                .group_by(FollowUp.customer_id)
            )
            for cid, last_dt in fu_last_result.all():
                last_follow_map[cid] = last_dt

        # Fetch customer sources
        from app.models.customer import CustomerSource
        source_ids = {c.source_id for c in customers if c.source_id}
        source_map = {}
        if source_ids:
            src_result = await self.db.execute(
                select(CustomerSource).where(CustomerSource.id.in_(source_ids))
            )
            for src in src_result.scalars().all():
                source_map[src.id] = src.name

        wb = Workbook()
        ws = wb.active
        ws.title = "客户报表"
        self._write_header(ws, CUSTOMER_HEADERS)

        for c in customers:
            ws.append([
                c.name,
                c.phone,
                CUSTOMER_STATUS_MAP.get(c.status.value if hasattr(c.status, "value") else str(c.status), str(c.status)),
                source_map.get(c.source_id, "") if c.source_id else "",
                c.budget_range or "",
                str(c.wedding_date) if c.wedding_date else "",
                user_map.get(c.assigned_sale_id, "") if c.assigned_sale_id else "",
                follow_up_map.get(c.id, 0),
                last_follow_map.get(c.id).strftime("%Y-%m-%d %H:%M:%S") if c.id in last_follow_map else "",
                c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
            ])

        return self._build_response(wb, "customer")

    async def export_finance(
        self,
        date_start: Optional[str] = None,
        date_end: Optional[str] = None,
        sale_id: Optional[int] = None,
    ) -> StreamingResponse:
        query = (
            select(Payment, Order.order_no, Order.total_amount, Order.paid_amount, Customer.name.label("customer_name"))
            .join(Order, Payment.order_id == Order.id)
            .join(Customer, Order.customer_id == Customer.id, isouter=True)
        )

        if date_start:
            query = query.where(Payment.created_at >= datetime.fromisoformat(date_start))
        if date_end:
            query = query.where(Payment.created_at <= datetime.fromisoformat(date_end))
        if sale_id:
            query = query.where(Order.sale_id == sale_id)

        query = query.order_by(Payment.created_at.desc()).limit(MAX_EXPORT_ROWS)
        result = await self.db.execute(query)
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "财务报表"
        self._write_header(ws, FINANCE_HEADERS)

        for row_data in rows:
            payment = row_data[0]
            order_no = row_data[1]
            total_amount = float(row_data[2])
            paid_amount = float(row_data[3])
            customer_name = row_data[4]
            pending = total_amount - paid_amount

            ws.append([
                order_no or "",
                customer_name or "",
                PAYMENT_METHOD_MAP.get(payment.method.value if hasattr(payment.method, "value") else str(payment.method), str(payment.method)),
                float(payment.amount),
                payment.paid_at.strftime("%Y-%m-%d %H:%M:%S") if payment.paid_at else "",
                total_amount,
                paid_amount,
                round(pending, 2),
            ])

        return self._build_response(wb, "finance")

    def _write_header(self, ws, headers: list[str]):
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)

    def _build_response(self, wb: Workbook, report_type: str) -> StreamingResponse:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"{report_type}_{timestamp}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
